from selenium import webdriver # pārlūka atvēršanai
from selenium.webdriver.chrome.service import Service # Chrome pārlūka serviss
from selenium.webdriver.common.by import By # atrod elementus
from selenium.webdriver.support.ui import Select # dropdown izvēlēm
import time # gaidīšanai
from datetime import date #uzzinat šodienas datumu
import pandas as pd  # datu apstrādei
import os #pareizo ceļu uz failu un bildes
import requests # attēlu lejupielādei
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import PatternFill
import uuid

sodiena=date.today().strftime("%d-%m-%Y")
service = Service()
option = webdriver.ChromeOptions()

print("============================")
print("=== SS.COM Aut Meklētājs ===")

print("ievadiet marku")
marka = input().strip().title()
if marka == "Bmw":
    marka = "BMW"
    print("ievadiet modeli ievērojot atstarpes un garumzīmes")
    modelis = input()
else:
    print("ievadiet modeli")
    modelis = input().strip().title()
print("ievadiet min tilpumu")
min_tilpums = input()
print("ievadiet max tilpumu")
max_tilpums = input()
print("ievadiet gadu no")
min_gads = input()
print("ievadiet gadu līdz")
max_gads = input()
max_cena = input("ievadiet maksimālo cenu")

driver = webdriver.Chrome(service=service, options=option)
url="https://www.ss.com/lv/transport/cars/"
driver.get(url)

find = driver.find_element(By.LINK_TEXT, marka) # atrod marku pec teksta
find.click()
time.sleep(1.5)

dropdowns = driver.find_elements(By.CLASS_NAME, 'filter_sel')
modelis_drop = Select(dropdowns[11]) # modelim nav koda vai id tadel atrod modeli pec dropdown indeksa
time.sleep(0.5)
modelis_drop.select_by_visible_text(modelis)
time.sleep(1.2)

min_tilp_drop = Select(driver.find_element(By.NAME, 'topt[15][min]'))
time.sleep(0.3)
min_tilp_drop.select_by_visible_text(min_tilpums)
time.sleep(0.5)
max_tilp_drop = Select(driver.find_element(By.NAME, 'topt[15][max]')) # atrod min un max laukus
time.sleep(0.6)
max_tilp_drop.select_by_visible_text(max_tilpums)
time.sleep(0.4)

min_gads_drop = Select(driver.find_element(By.NAME, 'topt[18][min]'))
time.sleep(0.1)
min_gads_drop.select_by_visible_text(min_gads)
time.sleep(0.2)
max_gads_drop = Select(driver.find_element(By.NAME, 'topt[18][max]'))
time.sleep(0.3)
max_gads_drop.select_by_visible_text(max_gads)
time.sleep(0.5)

max_cena_drop = driver.find_element(By.NAME, 'topt[8][max]')
max_cena_drop.send_keys(max_cena)
driver.find_element(By.XPATH, '//input[@type="submit" and contains(@class, "b")]').click()
time.sleep(1.5)

os.makedirs("bildes", exist_ok=True) # izveido mapi bildes 
pirms_vaksanas= driver.current_url
dati=[]
print(">>> Vāc rezultātus...")
print(">>> ---------------------")
while True:
    ads = driver.find_elements(By.CSS_SELECTOR, "tr[id^='tr_']")
    for ad in ads:
        try:
            title_element = ad.find_element(By.CLASS_NAME, "am")
            title = title_element.text
            link = title_element.get_attribute("href")
        except:
            title = "Nav norādīts nosaukums"
            link = "Nav norādīta saite"
        try:
            gads_element = ad.find_element(By.CLASS_NAME, "amopt")
            gads = gads_element.text
        except:
                gads = "Nav norādīta"
        try:
            tilpums_elements = ad.find_elements(By.CLASS_NAME, "amopt")
            if tilpums_elements:
                tilpums = tilpums_elements[1].text
        except:
                tilpums = "Nav norādīta"
        try:
            km_elements = ad.find_elements(By.CLASS_NAME, "amopt")
            if km_elements:
                km = km_elements[2].text
        except:
                km = "Nav norādīta"
        try:
            price_elements = ad.find_elements(By.CLASS_NAME, "amopt")
            if price_elements:
                price = price_elements[3].text
        except:
                price = "Nav norādīta cena"

        try:
            img = ad.find_element(By.TAG_NAME, "img")
            img_url = img.get_attribute("src")
            unikals_id = str(uuid.uuid4())[:8]
            img_path = f"bildes/{title[:30].replace(' ', '_')}_{unikals_id}.jpg"

            img_data = requests.get(img_url).content
            with open(img_path, "wb") as f:
                f.write(img_data)
        except:
            img_url = "Nav norādīts attēls"
            img_path = ""

            
        if title != "Nav norādīts nosaukums":
            print(f"Nosaukums: {title}")
            print(f"Saite: {link}")
            print(f"Gads: {gads}")
            print(f"Nobraukums: {km}")
            print(f"Tilpums: {tilpums}")
            print(f"Cena: {price}")
            print("------------------")
            dati.append({
                "Datums": sodiena,
                "Nosaukums": title,
                "Saite": link,
                "Gads": gads,
                "Tilpums": tilpums,
                "Nobraukums": km,
                "Cena": price,
                "Attēls": img_path
            })
    try:
        next_imgs = driver.find_elements(By.CSS_SELECTOR, 'a[rel="next"] img[src*="s_right.png"]')
        if next_imgs:
            next_button = next_imgs[0].find_element(By.XPATH, '..')

        # Ja nākamais URL ir tāds pats, ciklu pārtrauc

        driver.execute_script("arguments[0].scrollIntoView();", next_button)
        driver.execute_script("arguments[0].click();", next_button)
        time.sleep(1.5)
        current_url = driver.current_url
        if pirms_vaksanas == current_url:
            print(">>> Tika atvērta pirmā lapa vēlreiz — apstājamies.")
            break
    except:
        print(">>> Visi sludinājumi ir izskatīti!")
        break

input("\n>>> Nospied Enter, lai aizvērtu pārlūku...")
driver.quit()
print(">>> Pārlūks aizvērts.")

fails = pd.DataFrame(dati)

fails["score"] = 0.0
for i, row in fails.iterrows():
    try:
        gads = int(str(row["Gads"]).strip())
        km_text = str(row["Nobraukums"]).lower().replace(' ', '').replace('km', '').replace('tūkst.', '000').replace('tūkst', '000')
        if not km_text.isdigit():
            continue  # ja nobraukums nav zināms, izlaižam
        km = int(km_text)
        km_t = km / 1000 if km > 0 else 1
        score = gads / km_t
        fails.at[i, "score"] = score
    except Exception:
        fails.at[i, "score"] = 0.0

nosaukums= f"sludinajumi_{modelis}.xlsx"
uz_desk = os.path.join(os.path.expanduser("~"), f"Desktop", nosaukums)
if os.path.exists(uz_desk):
    esosais= pd.read_excel(uz_desk)
    kopejais= pd.concat([esosais, fails], ignore_index=True)
    kopejais.drop_duplicates(subset=["Saite"], inplace=True)
    kopejais.to_excel(uz_desk, index=False)
else:
    fails.to_excel(uz_desk, index=False)

# tabulas formatejums
wb = load_workbook(uz_desk)
ws = wb.active
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 45
ws.column_dimensions['C'].width = 60
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 10
ws.column_dimensions['F'].width = 15
ws.column_dimensions['G'].width = 12
ws.column_dimensions['H'].width = 18

fails_top3 = fails.sort_values(by="score", ascending=False).head(3)
top_links = set(fails_top3["Saite"].tolist())

for row_idx, row in enumerate(fails.itertuples(), start=2):
    img_file = getattr(row, "Attēls")
    link = getattr(row, "Saite")
    if os.path.exists(img_file) and img_file.endswith(".jpg"):
        img = XLImage(img_file)
        img.height = 80
        img.width = 80
        ws.add_image(img, f"H{row_idx}")
        ws.row_dimensions[row_idx].height = 60

    if link in top_links:
        for col in range(1, 9):
            ws.cell(row=row_idx, column=col).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

wb.save(uz_desk)
wb.close()

print(f">>> Dati saglabāti Excel failā: sludinajumi_{modelis}.xlsx")
os.startfile(uz_desk)
